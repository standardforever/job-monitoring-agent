from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from models.process import UploadDomainRow
from utils.logging import get_logger, log_event

logger = get_logger("file_input_service")


class FileInputService:
    def extract_upload_rows(self, filename: str, content: bytes) -> list[UploadDomainRow]:
        suffix = Path(filename).suffix.lower()
        log_event(
            logger,
            "info",
            "file_upload_row_extraction_started filename=%s suffix=%s",
            filename,
            suffix,
            domain=filename,
            upload_filename=filename,
            suffix=suffix,
        )

        if suffix == ".csv":
            rows = self._extract_rows_from_csv(content)
        elif suffix == ".xlsx":
            rows = self._extract_rows_from_xlsx(content)
        else:
            raise ValueError("Only .csv and .xlsx files are supported")

        if not rows:
            raise ValueError("No valid values found in the 'domain' column")

        log_event(
            logger,
            "info",
            "file_upload_row_extraction_completed filename=%s row_count=%s",
            filename,
            len(rows),
            domain=rows[0].domain if rows else filename,
            upload_filename=filename,
            row_count=len(rows),
        )
        return rows

    def extract_domains(self, filename: str, content: bytes) -> list[str]:
        return [row.domain for row in self.extract_upload_rows(filename, content)]

    def _extract_rows_from_csv(self, content: bytes) -> list[UploadDomainRow]:
        text_stream = StringIO(content.decode("utf-8-sig"))
        reader = csv.DictReader(text_stream)
        return self._collect_rows(reader)

    def _extract_rows_from_xlsx(self, content: bytes) -> list[UploadDomainRow]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        normalized_headers = [header.lower() for header in headers]
        if "domain" not in normalized_headers:
            raise ValueError("Uploaded file must contain a 'domain' column")

        domain_index = normalized_headers.index("domain")
        career_page_index = normalized_headers.index("career_page_url") if "career_page_url" in normalized_headers else None
        upload_rows: list[UploadDomainRow] = []
        seen: set[str] = set()
        for row in rows[1:]:
            if row is None or domain_index >= len(row):
                continue
            value = str(row[domain_index] or "").strip()
            if not value or value in seen:
                continue
            seen.add(value)
            career_page_url = None
            if career_page_index is not None and career_page_index < len(row):
                career_page_url = str(row[career_page_index] or "").strip() or None
            upload_rows.append(UploadDomainRow(domain=value, career_page_url=career_page_url))
        return upload_rows

    def _collect_rows(self, reader: csv.DictReader) -> list[UploadDomainRow]:
        if reader.fieldnames is None:
            return []

        normalized_fieldnames = {str(name).strip().lower(): name for name in reader.fieldnames if name}
        if "domain" not in normalized_fieldnames:
            raise ValueError("Uploaded file must contain a 'domain' column")

        domain_key = normalized_fieldnames["domain"]
        career_page_key = normalized_fieldnames.get("career_page_url")
        upload_rows: list[UploadDomainRow] = []
        seen: set[str] = set()
        for row in reader:
            value = str((row or {}).get(domain_key) or "").strip()
            if not value or value in seen:
                continue
            seen.add(value)
            career_page_url = str((row or {}).get(career_page_key) or "").strip() or None if career_page_key else None
            upload_rows.append(UploadDomainRow(domain=value, career_page_url=career_page_url))
        return upload_rows
